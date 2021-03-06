from urllib.request import urlopen

from bs4 import BeautifulSoup
import xlwt
from openpyxl import Workbook
from openpyxl import load_workbook
import re

NONE_VALUE = 'None value'
ROOT_WEBSITE = 'http://sklep.gandolf.pl/category.php?id_category=1820'
ROOT_CATEGORY_NAME = 'Gandolf'
XLS_FILENAME = 'gandolf_'
XLS_DESC = 'Opis'
XLS_CODE = 'Kod'
XLS_NAME = 'Nazwa'


class Product:
    name = NONE_VALUE
    desc = NONE_VALUE
    code = NONE_VALUE

    def __init__(self):
        super().__init__()
        self.name = NONE_VALUE
        self.desc = NONE_VALUE
        self.code = NONE_VALUE
        self.title = NONE_VALUE

    def print(self):
        if self.name:
            print('    Found product ---------------------------- ')
            print('    Product Name: ', self.name)
            print('    Product Desc: ', self.desc)
            print('    Product Code: ', self.code)
            print('    Product title: ', self.title)

    def is_defined(self):
        if self.name and self.desc and self.code and self.title:
            return True
        else:
            return False


class Category:

    def __init__(self, name):
        super().__init__()
        self.name = name
        self.products = []

    products = []
    name = NONE_VALUE

    def add(self, product):
        self.products.append(product)


def parse_category(soup, depth, category_name):
    categories_list = soup('ul', {'class': 'inline_list'})

    if not categories_list:
        parse_products(soup, category_name)
    else:
        for category in categories_list[0].find_all('a'):

            title = category.get('title')
            if title is not None:
                url_to_open = category.get('href') + '&n=100000'
                if depth is 0:
                    print('Found category: ', title)
                    category_soup = BeautifulSoup(urlopen(url_to_open).read())
                    parse_category(category_soup, depth + 1, title)
                elif depth is 1:
                    print('  Found subcategory: ',  title)
                    category_soup = BeautifulSoup(urlopen(url_to_open).read())
                    parse_category(category_soup, 1, category_name[:8] + '_' + title)


def parse_code(soup, product):
    code_soup = soup('p', {'id': 'product_reference'})
    if code_soup:
        code = code_soup[0].span.text
        product.code = code


def replace_with_newline(soup, element):
    for e in soup.findAll(element):
        e.replace_with('\n')


def parse_desc(soup, product):
    # replace_with_newline(soup, 'p')
    desc_soup = soup('div', {'id': 'idTab1'})
    if desc_soup:
        desc = desc_soup[0].text
        product.desc = desc


def parse_name(soup, product):
    name_soup = soup('div', {'id': 'primary_block'})
    if name_soup:
        name = name_soup[0].h1.text
        product.name = name


def parse_product(soup):
    replace_with_newline(soup, 'br')

    product = Product()
    parse_name(soup, product)
    parse_code(soup, product)
    parse_desc(soup, product)
    if product.is_defined():
        product.print()
        # products.append(product)
    return product


def parse_products(soup, name):

    category = Category(name)

    for link in soup('p', {'class': 'product_desc'}):
        title = link.a.get('title')

        if title is not None:
            url_to_open = link.a.get('href')
            product_soup = BeautifulSoup(urlopen(url_to_open).read())
            product = parse_product(product_soup)
            product.title = title
            category.add(product)

    categories.append(category)


def write_product_to_xls_sheet(index, product, sheet):
    sheet.cell(row=index + 2, column=1).value = product.code
    sheet.cell(row=index + 2, column=2).value = product.name
    sheet.cell(row=index + 2, column=3).value = product.title
    sheet.cell(row=index + 2, column=4).value = product.title[11:]


def write_category_to_xls_book(book, category):
    sheet = book.create_sheet()
    sheet.title = category.name[:31]
    sheet.cell(row=1, column=1).value = XLS_CODE
    sheet.cell(row=1, column=2).value = XLS_NAME
    sheet.cell(row=1, column=3).value = "Title"
    sheet.cell(row=1, column=4).value = "Short Title"
    return sheet


def write_to_xls():

    book = Workbook()

    for category in categories:

        sheet = write_category_to_xls_book(book, category)

        for index, product in enumerate(category.products):
            write_product_to_xls_sheet(index, product, sheet)

    book.save('gandolf.xlsx')

# --------------------------------------------------------------------------------------

categories = []

# soup = BeautifulSoup(urlopen(ROOT_WEBSITE).read())
# parse_category(soup, 0, ROOT_CATEGORY_NAME)
# write_to_xls()

# -------------------------------

book = Workbook()
wb = load_workbook(filename = 'gandolf.xlsx')
ws = book.active

i = 1
j = 1

for sheet in wb:
    for row in sheet.rows:
        ws.cell(row=i, column=1).value = row[0].value
        ws.cell(row=i, column=2).value = row[1].value
        ws.cell(row=i, column=3).value = row[2].value
        ws.cell(row=i, column=4).value = row[3].value
        i += 1

book.save('gandolf_jeden_arkusz.xlsx')

# ----------------------------------

# wb = load_workbook(filename='new2.xlsx')
#
# sheet = wb.active
#
# for row in sheet.rows:
#     if row[9].value is not None:
#         row[9].value = "<br />".join(row[9].value.split("\n"))
#         # row[9].value = re.sub(r'\s+', ' ', row[2].value)
#         # row[9].value = re.sub(r'\s+', ' ', row[2].value)
#
# wb.save('new3.xlsx')